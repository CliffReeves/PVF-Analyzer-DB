"""
rfq_parser.py — Auto-detecting parser for RFQ bid comparison spreadsheets.

Handles three observed formats:
  Format A: Bidder names in row 1+, column headers in a later row (e.g. row 5)
             Per-bidder cols: DELIVERY, WEEKS, UNIT PRICE, TOTAL PRICE, COMMENTS, MANUFACTURER
  Format B: Single header row 1, bidder name embedded as prefix: BIDDER_UNIT_PRICE
  Format C: Bidder names in row 1, column headers in row 2,
             Per-bidder cols: UNIT COST, EXT. COST, DELIVERY ARO, COMMENTS, DELIVERY DATE
"""

import re
import openpyxl


# ---------------------------------------------------------------------------
# Synonym maps — all lowercased for matching
# ---------------------------------------------------------------------------
UNIT_PRICE_SYNONYMS  = {"unit price", "unit cost", "unit_price", "unit_cost", "unitprice"}
EXT_PRICE_SYNONYMS   = {
    "total price", "ext. price", "ext price", "extended price",
    "ext. cost", "ext cost", "extended cost", "total_price",
    "ext_price", "ext_cost", "totalprice", "extprice"
}
ITEM_NUM_SYNONYMS    = {"item #", "item#", "item no", "item no.", "item number", "line no", "line no."}
DESC_SYNONYMS        = {"description", "desc"}
UNIT_SYNONYMS        = {"unit", "units", "unit of measure", "uom"}
QTY_SYNONYMS         = {"qty quoted", "qty", "quantity", "units quoted", "qty quoted"}
SIZE_SYNONYMS        = {"size"}


def _norm(val):
    """Normalise a cell value to a lower-stripped string, or '' if None."""
    if val is None:
        return ""
    return str(val).strip().lower()


def _matches(val, synonyms):
    return _norm(val) in synonyms


def _find_header_row(rows):
    """
    Return (header_row_index, row_values_list) for the row that looks like
    a column-header row (contains an item# AND description-like column).
    Searches the first 15 rows.
    """
    for i, row in enumerate(rows[:15]):
        has_item = any(_matches(c, ITEM_NUM_SYNONYMS) for c in row)
        has_desc = any(_matches(c, DESC_SYNONYMS) for c in row)
        if has_item and has_desc:
            return i, list(row)
    return None, None


def _detect_format_b(header_cols):
    """
    Format B: any column header matches BIDDER_UNIT_PRICE or BIDDER_TOTAL_PRICE pattern.
    Handles both underscore-separated (WHITCO_UNIT_PRICE) and
    space-separated (WHITCO UNIT PRICE) variants, including EXT> PRICE typo.
    """
    # Underscore-separated detection (original)
    for h in header_cols:
        s = _norm(h)
        if s and "_" in s:
            parts = s.rsplit("_", 1)
            if parts[-1] in {"unit_price", "unit price", "total_price", "total price",
                              "unit_cost", "ext_price", "ext_cost", "unit cost", "ext cost"}:
                return True
            if re.search(r"(unit.?price|unit.?cost|total.?price|ext.?price|ext.?cost)", s):
                return True
    for h in header_cols:
        s = str(h).strip() if h else ""
        if re.match(r"^[A-Z][A-Z0-9\-]+_(UNIT_PRICE|TOTAL_PRICE|UNIT_COST|EXT_COST|EXT_PRICE)",
                    s, re.IGNORECASE):
            return True
    # Space-separated detection: "BIDDER UNIT PRICE", "BIDDER EXT. PRICE", "BIDDER EXT> PRICE"
    _PRICE_ENDINGS_2W = {
        "unit price", "unit cost", "total price",
        "ext price", "ext. price", "extended price",
        "ext cost", "ext. cost", "extended cost",
    }
    for h in header_cols:
        if not h:
            continue
        words = str(h).strip().lower().replace(">", ".").split()
        if len(words) >= 3:
            last_two = " ".join(words[-2:])
            if last_two in _PRICE_ENDINGS_2W:
                return True
    return False


def _parse_format_b(rows, header_idx, header_cols):
    """
    Parse Format B: single header row with BIDDER_FIELD columns.
    Returns (rfq_cols_map, bidder_map).
    rfq_cols_map: {field_name: col_index}
    bidder_map: {bidder_name: {unit_price: col, ext_price: col}}
    """
    rfq_map = {}
    bidder_map = {}

    for col_i, h in enumerate(header_cols):
        if h is None:
            continue
        s = str(h).strip()
        s_low = s.lower()

        # Check if it's a compound BIDDER_FIELD column
        # Try to split on the last underscore-separated keyword
        matched_bidder = None
        matched_field  = None

        for syn in UNIT_PRICE_SYNONYMS:
            syn_under = syn.replace(" ", "_")
            if s_low.endswith("_" + syn_under) or s_low.endswith("_" + syn):
                candidate = s[:-(len(syn_under)+1)].rstrip("_")
                matched_bidder = candidate.upper()
                matched_field  = "unit_price"
                break
        if not matched_bidder:
            for syn in EXT_PRICE_SYNONYMS:
                syn_under = syn.replace(" ", "_").replace(".", "")
                s_test = s_low.replace(".", "").replace(" ", "_")
                syn_test = syn_under.replace(".", "")
                if s_test.endswith("_" + syn_test):
                    candidate = s[:-(len(syn_under)+1)].rstrip("_")
                    matched_bidder = candidate.upper()
                    matched_field  = "ext_price"
                    break

        # Fallback: regex split on well-known underscore patterns
        if not matched_bidder:
            m = re.match(
                r"^(.+?)_(UNIT.?PRICE|UNIT.?COST|TOTAL.?PRICE|EXT.?PRICE|EXT.?COST)$",
                s, re.IGNORECASE
            )
            if m:
                matched_bidder = m.group(1).upper()
                field_raw = m.group(2).lower()
                if any(x in field_raw for x in ["unit", "cost"]) and "total" not in field_raw and "ext" not in field_raw:
                    matched_field = "unit_price"
                else:
                    matched_field = "ext_price"

        # Fallback: space-separated "BIDDER UNIT PRICE" / "BIDDER EXT. PRICE"
        # Also handles EXT> PRICE typo (> instead of .)
        if not matched_bidder:
            words = s.replace(">", ".").split()
            if len(words) >= 3:
                last_two = " ".join(words[-2:]).lower()
                if last_two in {"unit price", "unit cost"}:
                    matched_bidder = words[0].upper()
                    matched_field  = "unit_price"
                elif last_two in {"total price", "ext price", "ext. price",
                                  "extended price", "ext cost", "ext. cost", "extended cost"}:
                    matched_bidder = words[0].upper()
                    matched_field  = "ext_price"

        if matched_bidder and matched_field:
            if matched_bidder not in bidder_map:
                bidder_map[matched_bidder] = {}
            bidder_map[matched_bidder][matched_field] = col_i
            continue

        # RFQ columns
        if _matches(h, ITEM_NUM_SYNONYMS):
            rfq_map["item_num"] = col_i
        elif _matches(h, DESC_SYNONYMS):
            rfq_map["description"] = col_i
        elif _matches(h, UNIT_SYNONYMS):
            rfq_map["unit"] = col_i
        elif _matches(h, QTY_SYNONYMS):
            rfq_map["quantity"] = col_i
        elif _matches(h, SIZE_SYNONYMS):
            rfq_map["size"] = col_i

    return rfq_map, bidder_map


def _find_bidder_names_above(rows, header_idx):
    """
    Scan rows 0..header_idx-1 looking for a row that contains recognisable
    bidder name tokens. Returns dict {col_index: bidder_name}.

    Strategy: prefer the TOP-MOST qualifying row (closest to row 0), because
    company names (SMP, EDGEN, WHITCO) appear in the first row while contact
    details (names, phones, emails) appear in subsequent rows.

    A qualifying row has >= 2 cells that are:
      - not None / empty
      - not phone-like or email-like
      - non-purely-numeric
      - short enough to be a company abbreviation/name (< 40 chars)
      - not a known skip-word (RFQ, DELIVERY, WEEKS, etc.)
    """
    SKIP_WORDS = {"rfq", "none", "delivery", "weeks", "manufacturer",
                  "comments", "vendor comments", "unit price", "total price"}

    best_row   = {}
    best_score = 0   # prefer rows with shorter average candidate length (abbreviations)

    for i in range(header_idx):
        row = rows[i]
        candidates = {}
        for col_i, val in enumerate(row):
            if val is None:
                continue
            s = str(val).strip()
            if not s or len(s) > 50:
                continue
            if _norm(s) in SKIP_WORDS:
                continue
            # Skip phone-like values
            if re.match(r"^[\d\s\.\-\(\)\+]+$", s):
                continue
            # Skip email addresses
            if "@" in s:
                continue
            # Skip long multi-word strings that look like contact names or addresses
            words = s.split()
            if len(words) > 4:
                continue
            candidates[col_i] = s.upper()

        if len(candidates) >= 2:
            avg_len = sum(len(v) for v in candidates.values()) / len(candidates)
            # Prefer shorter names (company abbrevs) over longer names (full contact names)
            score = 1.0 / avg_len
            if score > best_score:
                best_score = score
                best_row   = candidates

    return best_row


def _parse_format_ac(rows, header_idx, header_cols):
    """
    Parse Format A and C: bidder names in a row above the header.
    Detect RFQ columns and per-bidder column groups.
    Returns (rfq_cols_map, bidder_map).
    """
    # Step 1: find bidder names above header
    bidder_names_at_col = _find_bidder_names_above(rows, header_idx)

    # Step 2: map header columns
    rfq_map   = {}
    col_roles  = {}  # col_i -> "unit_price" | "ext_price" | None

    for col_i, h in enumerate(header_cols):
        if h is None:
            continue
        if _matches(h, ITEM_NUM_SYNONYMS):
            rfq_map["item_num"] = col_i
        elif _matches(h, DESC_SYNONYMS):
            rfq_map["description"] = col_i
        elif _matches(h, UNIT_SYNONYMS):
            rfq_map["unit"] = col_i
        elif _matches(h, QTY_SYNONYMS):
            rfq_map["quantity"] = col_i
        elif _matches(h, SIZE_SYNONYMS):
            rfq_map["size"] = col_i
        elif _norm(h) in UNIT_PRICE_SYNONYMS:
            col_roles[col_i] = "unit_price"
        elif _norm(h) in EXT_PRICE_SYNONYMS:
            col_roles[col_i] = "ext_price"

    # Step 3: assign price columns to the nearest bidder on the left
    bidder_start_cols = sorted(bidder_names_at_col.keys())

    def _bidder_for_col(col_i):
        owner = None
        for bc in bidder_start_cols:
            if bc <= col_i:
                owner = bc
            else:
                break
        if owner is not None:
            return bidder_names_at_col[owner]
        return None

    bidder_map = {}
    for col_i, role in col_roles.items():
        name = _bidder_for_col(col_i)
        if name:
            if name not in bidder_map:
                bidder_map[name] = {}
            bidder_map[name][role] = col_i

    return rfq_map, bidder_map


def _extract_type_spec(description, size_val=None):
    """
    Split description into (item_type, specification).
    The first word is the item type; the rest is the specification.
    If size_val is provided and not in spec, prepend it.
    """
    if not description:
        return "", ""
    description = str(description).strip()
    # Split on first comma
    if "," in description:
        parts = description.split(",", 1)
        item_type = parts[0].strip().upper()
        spec      = parts[1].strip()
    else:
        # Split on first space
        parts = description.split(None, 1)
        item_type = parts[0].strip().upper() if parts else ""
        spec      = parts[1].strip() if len(parts) > 1 else ""

    return item_type, spec


def _is_data_row(row, rfq_map):
    """Return True if this row looks like an item data row (has an item number)."""
    item_col = rfq_map.get("item_num")
    if item_col is None or item_col >= len(row):
        return False
    val = row[item_col]
    if val is None:
        return False
    s = str(val).strip()
    # Item numbers are typically numeric or alphanumeric short codes
    return bool(s) and len(s) < 20


def _safe_float(val):
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        s = str(val).replace(",", "").replace("$", "").strip()
        try:
            return float(s)
        except (ValueError, TypeError):
            return None


def parse_excel(filepath, sheet_name=None):
    """
    Main entry point. Parse an RFQ Excel file.

    Returns a dict:
    {
      "format": "A" | "B" | "C",
      "sheet": sheet_name,
      "bidders": [list of bidder name strings],
      "items": [
        {
          "item_number": str,
          "item_type": str,
          "specification": str,
          "size": str | None,
          "unit": str | None,
          "quantity": float | None,
          "bids": {
            "BIDDER_NAME": {"unit_price": float|None, "ext_price": float|None},
            ...
          }
        },
        ...
      ],
      "ambiguities": [list of warning strings]
    }
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Pick the best sheet
    if sheet_name:
        ws = wb[sheet_name]
    else:
        # Prefer sheets with the most data; skip Dashboard/Terms sheets
        best = None
        best_rows = 0
        skip_keywords = {"dashboard", "terms", "condition", "sheet2", "alternative"}
        for sn in wb.sheetnames:
            if any(k in sn.lower() for k in skip_keywords):
                continue
            s = wb[sn]
            if s.max_row > best_rows:
                best_rows = s.max_row
                best = sn
        if best is None:
            best = wb.sheetnames[0]
        ws = wb[best]
        sheet_name = best

    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    if not rows:
        return {"error": "Sheet is empty", "sheet": sheet_name}

    # Find header row
    header_idx, header_cols = _find_header_row(rows)
    if header_idx is None:
        return {"error": "Could not locate column header row", "sheet": sheet_name}

    # Detect format
    if _detect_format_b(header_cols):
        fmt = "B"
        rfq_map, bidder_map = _parse_format_b(rows, header_idx, header_cols)
    else:
        fmt = "AC"
        rfq_map, bidder_map = _parse_format_ac(rows, header_idx, header_cols)

    ambiguities = []
    if not bidder_map:
        ambiguities.append("No bidders detected automatically.")
    if "item_num" not in rfq_map:
        ambiguities.append("Could not locate Item # column.")
    if "description" not in rfq_map:
        ambiguities.append("Could not locate Description column.")

    # Parse data rows
    items = []
    for row in rows[header_idx + 1:]:
        if not _is_data_row(row, rfq_map):
            continue

        def _cell(col_key):
            col = rfq_map.get(col_key)
            if col is not None and col < len(row):
                return row[col]
            return None

        raw_desc = _cell("description")
        raw_size = _cell("size")
        raw_unit = _cell("unit")
        raw_qty  = _cell("quantity")
        raw_item = _cell("item_num")

        item_type, spec = _extract_type_spec(raw_desc, raw_size)

        bids = {}
        for bidder_name, cols in bidder_map.items():
            up_col = cols.get("unit_price")
            ep_col = cols.get("ext_price")
            up_val = _safe_float(row[up_col]) if (up_col is not None and up_col < len(row)) else None
            ep_val = _safe_float(row[ep_col]) if (ep_col is not None and ep_col < len(row)) else None
            bids[bidder_name] = {"unit_price": up_val, "ext_price": ep_val}

        items.append({
            "item_number":   str(raw_item).strip() if raw_item else "",
            "item_type":     item_type,
            "specification": spec,
            "size":          str(raw_size).strip() if raw_size else None,
            "unit":          str(raw_unit).strip().upper() if raw_unit else None,
            "quantity":      _safe_float(raw_qty),
            "bids":          bids,
        })

    return {
        "format":       fmt,
        "sheet":        sheet_name,
        "bidders":      sorted(bidder_map.keys()),
        "items":        items,
        "ambiguities":  ambiguities,
        "rfq_map":      rfq_map,
    }


def list_sheets(filepath):
    """Return list of sheet names in the workbook."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    return wb.sheetnames
